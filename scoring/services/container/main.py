"""Container service: the one agent monitoring the 1B-MNT fill.

Endpoints
─────────
    GET  /healthz
    GET  /container                     — frozen target + live fill
    GET  /container/inflows             — recent inflows (cursor on UUID7 id)
    POST /container/inflows             — record one inflow
    GET  /milestones                    — frozen MILESTONES + per-id progress
    POST /milestones/{id}/status        — update progress for one milestone
    GET  /scenarios                     — frozen SCENARIOS catalog
    POST /scenarios/simulate            — run one of bear/base/bull
    GET  /scenarios/runs                — recent simulation runs
    POST /excel/import                  — upload XLSX of inflows
    GET  /excel/export                  — download XLSX of current state
    GET  /agent/report                  — single-agent monitor report

All ids are UUID7. Inflows append-only. Constants from shared.constants
are NEVER mutated by this service — only read.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import sqlite3
from datetime import datetime, timezone
from typing import Literal, Optional

from fastapi import Depends, FastAPI, File, HTTPException, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from shared import witness
from shared.constants import (
    CHANNELS, CONTAINER_AGENT_HANDLE, CONTAINER_CURRENCY, CONTAINER_END,
    CONTAINER_START, CONTAINER_TARGET, MILESTONES, SCENARIOS,
    milestone_by_id, total_expected_revenue,
)
from shared.db import new_id, transaction
from shared.deps import require_admin, require_session

CORS_ORIGIN = os.environ.get("CORS_ORIGIN", "http://127.0.0.1:8010")

app = FastAPI(title="scoring-container", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[CORS_ORIGIN],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)
SERVICE = "container"


# ─────────────────────────────────────── pydantic models

class InflowIn(BaseModel):
    occurred_at: str
    amount_mnt: float
    channel: str
    milestone_id: Optional[str] = None
    note: Optional[str] = None
    source: Literal["manual", "excel", "api", "agent"] = "manual"
    actor_id: Optional[str] = None


class MilestoneStatusIn(BaseModel):
    status: Literal["pending", "in_progress", "done", "blocked", "missed"]
    completion_pct: float = Field(ge=0, le=100, default=0)
    notes: Optional[str] = None
    actor_id: Optional[str] = None


class ScenarioRunIn(BaseModel):
    scenario_key: Literal["bear", "base", "bull"]
    actor_id: Optional[str] = None


# ─────────────────────────────────────── helpers

def _row(r: sqlite3.Row | None) -> dict | None:
    return dict(r) if r else None


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _container_state(conn) -> dict:
    """Live container state: target + filled + remaining + pace."""
    target = CONTAINER_TARGET
    row = conn.execute(
        "SELECT COALESCE(SUM(amount_mnt), 0) AS filled FROM container_inflows"
    ).fetchone()
    filled = float(row["filled"] or 0)
    remaining = max(0.0, target - filled)
    pct = (filled / target * 100.0) if target else 0.0

    # Pace: days elapsed vs total days. We use UTC dates from constants.
    start = datetime.fromisoformat(CONTAINER_START + "T00:00:00+00:00")
    end = datetime.fromisoformat(CONTAINER_END + "T23:59:59+00:00")
    now = datetime.now(timezone.utc)
    total_days = max(1, (end - start).days)
    elapsed_days = max(0, (now - start).days)
    elapsed_days = min(total_days, elapsed_days)
    expected_pct_by_pace = elapsed_days / total_days * 100.0

    inflow_count = conn.execute(
        "SELECT COUNT(*) AS c FROM container_inflows"
    ).fetchone()["c"]

    return {
        "target_mnt": target,
        "currency": CONTAINER_CURRENCY,
        "started_at": CONTAINER_START,
        "ends_at": CONTAINER_END,
        "filled_mnt": filled,
        "remaining_mnt": remaining,
        "fill_pct": round(pct, 2),
        "expected_pct_by_pace": round(expected_pct_by_pace, 2),
        "ahead_or_behind_pct": round(pct - expected_pct_by_pace, 2),
        "elapsed_days": elapsed_days,
        "total_days": total_days,
        "inflow_count": inflow_count,
    }


@app.get("/healthz")
def healthz() -> dict:
    return {"service": SERVICE, "ok": True}


# ─────────────────────────────────────── container state

@app.get("/container")
def get_container() -> dict:
    with transaction() as conn:
        state = _container_state(conn)
        agent = conn.execute(
            "SELECT id, name, handle FROM team_workers WHERE handle=?",
            (CONTAINER_AGENT_HANDLE,),
        ).fetchone()
        state["monitor"] = _row(agent)
        return state


@app.get("/container/inflows")
def list_inflows(limit: int = 50, before: str | None = None) -> dict:
    limit = max(1, min(limit, 200))
    with transaction() as conn:
        if before:
            rows = conn.execute(
                """SELECT * FROM container_inflows WHERE id < ?
                   ORDER BY id DESC LIMIT ?""",
                (before, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM container_inflows ORDER BY id DESC LIMIT ?",
                (limit,),
            ).fetchall()
        return {"inflows": [_row(r) for r in rows]}


@app.post("/container/inflows", status_code=201)
def add_inflow(body: InflowIn, session: dict = Depends(require_session)) -> dict:
    if body.channel not in CHANNELS:
        raise HTTPException(422, f"unknown channel: {body.channel}; allowed: {CHANNELS}")
    if body.milestone_id and milestone_by_id(body.milestone_id) is None:
        raise HTTPException(422, f"unknown milestone_id: {body.milestone_id}")

    actor = body.actor_id or session.get("worker_id")
    with transaction() as conn:
        fid = new_id()
        conn.execute(
            """INSERT INTO container_inflows
               (id, occurred_at, amount_mnt, channel, milestone_id, note, source, recorded_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (fid, body.occurred_at, body.amount_mnt, body.channel,
             body.milestone_id, body.note, body.source, actor),
        )

        # If this inflow ties to a milestone, increment its actual_revenue.
        if body.milestone_id:
            conn.execute(
                """UPDATE container_milestone_progress
                   SET actual_revenue_mnt = actual_revenue_mnt + ?,
                       updated_at = datetime('now'),
                       updated_by = ?
                   WHERE milestone_id = ?""",
                (body.amount_mnt, actor, body.milestone_id),
            )

        witness.append(
            conn,
            actor_id=actor,
            service=SERVICE,
            action="inflow.recorded",
            target_type="inflow",
            target_id=fid,
            payload={"amount_mnt": body.amount_mnt, "channel": body.channel,
                     "milestone_id": body.milestone_id, "source": body.source},
        )
        return _row(conn.execute(
            "SELECT * FROM container_inflows WHERE id=?", (fid,)
        ).fetchone())


# ─────────────────────────────────────── milestones (CONSTANT defs + DB progress)

@app.get("/milestones")
def list_milestones() -> dict:
    """Returns the frozen 27 milestones from constants.py joined with their
    runtime progress rows. The plan itself is read from code, not the DB."""
    with transaction() as conn:
        progress = {
            r["milestone_id"]: _row(r)
            for r in conn.execute("SELECT * FROM container_milestone_progress").fetchall()
        }
    out = []
    for m in MILESTONES:
        p = progress.get(m["id"], {})
        out.append({
            "id":             m["id"],
            "date":           m["date"],
            "track":          m["track"],
            "phase":          m["phase"],
            "title":          m["title"],
            "dod":            m["dod"],
            "kpi":            m["kpi"],
            "owners":         list(m["owners"]),
            "critical":       m["critical"],
            "expected_revenue_mnt": m["expected_revenue_mnt"],
            "status":         p.get("status", "pending"),
            "completion_pct": p.get("completion_pct", 0),
            "actual_revenue_mnt": p.get("actual_revenue_mnt", 0),
            "completed_at":   p.get("completed_at"),
            "notes":          p.get("notes"),
        })
    return {
        "milestones": out,
        "total_count": len(out),
        "critical_count": sum(1 for m in MILESTONES if m["critical"]),
        "total_expected_revenue_mnt": total_expected_revenue(),
    }


@app.post("/milestones/{milestone_id}/status")
def update_milestone(
    milestone_id: str,
    body: MilestoneStatusIn,
    session: dict = Depends(require_session),
) -> dict:
    if milestone_by_id(milestone_id) is None:
        raise HTTPException(404, f"unknown milestone: {milestone_id}")
    actor = body.actor_id or session.get("worker_id")
    completed_at = _now_iso() if body.status == "done" else None
    with transaction() as conn:
        conn.execute(
            """UPDATE container_milestone_progress
               SET status=?, completion_pct=?, notes=?, completed_at=COALESCE(?, completed_at),
                   updated_at=datetime('now'), updated_by=?
               WHERE milestone_id=?""",
            (body.status, body.completion_pct, body.notes, completed_at, actor, milestone_id),
        )
        witness.append(
            conn,
            actor_id=actor,
            service=SERVICE,
            action="milestone.status",
            target_type="milestone",
            target_id=milestone_id,
            payload={"status": body.status, "completion_pct": body.completion_pct},
        )
        row = conn.execute(
            "SELECT * FROM container_milestone_progress WHERE milestone_id=?",
            (milestone_id,),
        ).fetchone()
        return _row(row)


# ─────────────────────────────────────── 3-scenario simulation

def _simulate(scenario_key: str) -> dict:
    """Pure function: read constants, compute scenario output. No DB."""
    sc = SCENARIOS[scenario_key]
    channels_out = []
    total_revenue = 0.0
    total_gross_profit = 0.0
    total_orders = 0
    for ch_name, ch in sc["channels"].items():
        revenue = ch["orders"] * ch["aov_mnt"]
        gross_profit = revenue * ch["gross_margin"]
        channels_out.append({
            "channel": ch_name,
            "orders": ch["orders"],
            "aov_mnt": ch["aov_mnt"],
            "gross_margin": ch["gross_margin"],
            "revenue_mnt": revenue,
            "gross_profit_mnt": gross_profit,
        })
        total_revenue += revenue
        total_gross_profit += gross_profit
        total_orders += ch["orders"]

    marketing = sc["marketing_budget_mnt"]
    operating = sc["operating_expense_mnt"]
    a_commission_pct = sc["a_commission_pct"]
    a_commission = total_revenue * a_commission_pct

    # B-side net profit after all OpEx + A-side commission
    b_net = total_gross_profit - marketing - operating - a_commission
    # A-side net = commission - assumed labor cost of A team (~35M)
    a_net = a_commission - 35_000_000

    return {
        "scenario_key": scenario_key,
        "label": sc["label"],
        "channels": channels_out,
        "total_orders": total_orders,
        "total_revenue_mnt": total_revenue,
        "total_gross_profit_mnt": total_gross_profit,
        "marketing_budget_mnt": marketing,
        "operating_expense_mnt": operating,
        "a_commission_pct": a_commission_pct,
        "a_commission_mnt": a_commission,
        "a_net_profit_mnt": a_net,
        "b_net_profit_mnt": b_net,
        "goal_pct": round(total_revenue / CONTAINER_TARGET * 100.0, 2),
        "win_win": a_net > 0 and b_net > 0,
    }


@app.get("/scenarios")
def list_scenarios() -> dict:
    return {
        "scenarios": {k: {"label": v["label"], "channels": list(v["channels"].keys())}
                      for k, v in SCENARIOS.items()},
        "target_mnt": CONTAINER_TARGET,
    }


@app.post("/scenarios/simulate", status_code=201)
def simulate(body: ScenarioRunIn, session: dict = Depends(require_session)) -> dict:
    if body.scenario_key not in SCENARIOS:
        raise HTTPException(422, f"unknown scenario: {body.scenario_key}")
    actor = body.actor_id or session.get("worker_id")
    result = _simulate(body.scenario_key)
    with transaction() as conn:
        rid = new_id()
        conn.execute(
            """INSERT INTO container_scenario_runs
               (id, scenario_key, inputs_json, output_json,
                total_revenue_mnt, goal_pct, a_profit_mnt, b_profit_mnt, run_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (rid, body.scenario_key, json.dumps(SCENARIOS[body.scenario_key]),
             json.dumps(result, default=float),
             result["total_revenue_mnt"], result["goal_pct"],
             result["a_net_profit_mnt"], result["b_net_profit_mnt"], actor),
        )
        witness.append(
            conn,
            actor_id=actor,
            service=SERVICE,
            action="scenario.simulated",
            target_type="scenario",
            target_id=rid,
            payload={"scenario": body.scenario_key,
                     "revenue": result["total_revenue_mnt"],
                     "goal_pct": result["goal_pct"]},
        )
    result["run_id"] = rid
    return result


@app.get("/scenarios/runs")
def list_runs(limit: int = 20) -> dict:
    limit = max(1, min(limit, 100))
    with transaction() as conn:
        rows = conn.execute(
            """SELECT id, scenario_key, total_revenue_mnt, goal_pct,
                      a_profit_mnt, b_profit_mnt, run_at, run_by
               FROM container_scenario_runs ORDER BY id DESC LIMIT ?""",
            (limit,),
        ).fetchall()
        return {"runs": [_row(r) for r in rows]}


# ─────────────────────────────────────── Excel import / export

def _import_xlsx(content: bytes) -> tuple[list[dict], list[str]]:
    """Parse XLSX → list of inflow dicts. Returns (rows, errors).
    Expected columns (case-insensitive, first sheet):
        occurred_at | amount_mnt | channel | milestone_id (optional) | note (optional)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server")

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return [], ["empty file"]
    headers = [str(c).strip().lower() if c else "" for c in header_row]

    def col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    i_occurred = col("occurred_at")
    i_amount = col("amount_mnt")
    i_channel = col("channel")
    i_milestone = col("milestone_id")
    i_note = col("note")

    if min(i_occurred, i_amount, i_channel) < 0:
        return [], ["missing required column: occurred_at / amount_mnt / channel"]

    rows: list[dict] = []
    errors: list[str] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        try:
            occurred = row[i_occurred]
            if isinstance(occurred, datetime):
                occurred_at = occurred.isoformat()
            else:
                occurred_at = str(occurred).strip()
            amount = float(row[i_amount])
            channel = str(row[i_channel]).strip()
            if channel not in CHANNELS:
                errors.append(f"row {ridx}: unknown channel {channel!r}")
                continue
            milestone_id = None
            if i_milestone >= 0 and row[i_milestone]:
                mid = str(row[i_milestone]).strip()
                if milestone_by_id(mid):
                    milestone_id = mid
                else:
                    errors.append(f"row {ridx}: unknown milestone_id {mid!r}")
                    continue
            note = None
            if i_note >= 0 and row[i_note]:
                note = str(row[i_note])
            rows.append({
                "occurred_at": occurred_at,
                "amount_mnt": amount,
                "channel": channel,
                "milestone_id": milestone_id,
                "note": note,
            })
        except Exception as e:
            errors.append(f"row {ridx}: {e}")
    return rows, errors


@app.post("/excel/import")
async def excel_import(
    file: UploadFile = File(...),
    session: dict = Depends(require_admin),
) -> dict:
    content = await file.read()
    sha = hashlib.sha256(content).hexdigest()
    rows, errors = _import_xlsx(content)

    actor = session.get("worker_id")
    inserted = 0
    with transaction() as conn:
        for r in rows:
            fid = new_id()
            conn.execute(
                """INSERT INTO container_inflows
                   (id, occurred_at, amount_mnt, channel, milestone_id, note, source, recorded_by)
                   VALUES (?, ?, ?, ?, ?, ?, 'excel', ?)""",
                (fid, r["occurred_at"], r["amount_mnt"], r["channel"],
                 r["milestone_id"], r["note"], actor),
            )
            if r["milestone_id"]:
                conn.execute(
                    """UPDATE container_milestone_progress
                       SET actual_revenue_mnt = actual_revenue_mnt + ?
                       WHERE milestone_id = ?""",
                    (r["amount_mnt"], r["milestone_id"]),
                )
            inserted += 1

        sync_id = new_id()
        conn.execute(
            """INSERT INTO container_excel_sync
               (id, direction, rows_count, bytes, sha256, note, actor_id)
               VALUES (?, 'import', ?, ?, ?, ?, ?)""",
            (sync_id, inserted, len(content), sha,
             f"{len(errors)} errors" if errors else None, actor),
        )
        witness.append(
            conn,
            actor_id=actor,
            service=SERVICE,
            action="excel.imported",
            target_type="excel_sync",
            target_id=sync_id,
            payload={"rows": inserted, "errors": errors[:10], "sha256": sha},
        )
    return {
        "imported": inserted,
        "errors": errors,
        "sha256": sha,
        "bytes": len(content),
    }


@app.get("/excel/export")
def excel_export(session: dict = Depends(require_session)) -> StreamingResponse:
    """Export current state (container summary + milestones + inflows + last run) to XLSX."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server")

    wb = Workbook()
    BOLD = Font(bold=True)
    HEAD = PatternFill("solid", start_color="1F4E78")
    HEAD_FONT = Font(color="FFFFFF", bold=True)
    CENTER = Alignment(horizontal="center")

    with transaction() as conn:
        state = _container_state(conn)
        milestones_rows = conn.execute(
            "SELECT * FROM container_milestone_progress ORDER BY milestone_id"
        ).fetchall()
        progress = {r["milestone_id"]: _row(r) for r in milestones_rows}
        inflows = conn.execute(
            """SELECT id, occurred_at, amount_mnt, channel, milestone_id, note, source, recorded_at
               FROM container_inflows ORDER BY occurred_at"""
        ).fetchall()
        last_run = conn.execute(
            "SELECT * FROM container_scenario_runs ORDER BY id DESC LIMIT 3"
        ).fetchall()

    # Sheet 1: Container
    ws = wb.active
    ws.title = "Container"
    rows = [
        ("Target MNT", state["target_mnt"]),
        ("Filled MNT", state["filled_mnt"]),
        ("Remaining MNT", state["remaining_mnt"]),
        ("Fill %", state["fill_pct"]),
        ("Expected % by pace", state["expected_pct_by_pace"]),
        ("Ahead/behind %", state["ahead_or_behind_pct"]),
        ("Elapsed days", state["elapsed_days"]),
        ("Total days", state["total_days"]),
        ("Inflow count", state["inflow_count"]),
        ("Started at", state["started_at"]),
        ("Ends at", state["ends_at"]),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20

    # Sheet 2: Milestones
    ws2 = wb.create_sheet("Milestones")
    headers = ["ID", "Date", "Track", "Phase", "Title", "Critical",
               "Expected MNT", "Status", "Completion %", "Actual MNT", "Owners"]
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD
        c.alignment = CENTER
    for r, m in enumerate(MILESTONES, start=2):
        p = progress.get(m["id"], {})
        ws2.cell(row=r, column=1, value=m["id"])
        ws2.cell(row=r, column=2, value=m["date"])
        ws2.cell(row=r, column=3, value=m["track"])
        ws2.cell(row=r, column=4, value=m["phase"])
        ws2.cell(row=r, column=5, value=m["title"])
        ws2.cell(row=r, column=6, value="★" if m["critical"] else "")
        ws2.cell(row=r, column=7, value=m["expected_revenue_mnt"])
        ws2.cell(row=r, column=8, value=p.get("status", "pending"))
        ws2.cell(row=r, column=9, value=p.get("completion_pct", 0))
        ws2.cell(row=r, column=10, value=p.get("actual_revenue_mnt", 0))
        ws2.cell(row=r, column=11, value=", ".join(m["owners"]))
    for col, w in zip("ABCDEFGHIJK", [6, 12, 8, 7, 38, 9, 14, 12, 12, 14, 28]):
        ws2.column_dimensions[col].width = w

    # Sheet 3: Inflows
    ws3 = wb.create_sheet("Inflows")
    headers = ["occurred_at", "amount_mnt", "channel", "milestone_id", "note", "source", "recorded_at"]
    for i, h in enumerate(headers, start=1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD
    for r, row in enumerate(inflows, start=2):
        for i, k in enumerate(headers, start=1):
            ws3.cell(row=r, column=i, value=row[k])
    for col, w in zip("ABCDEFG", [22, 16, 24, 14, 30, 10, 22]):
        ws3.column_dimensions[col].width = w

    # Sheet 4: Scenario runs
    ws4 = wb.create_sheet("Scenario_Runs")
    headers = ["id", "scenario_key", "total_revenue_mnt", "goal_pct", "a_profit_mnt", "b_profit_mnt", "run_at"]
    for i, h in enumerate(headers, start=1):
        c = ws4.cell(row=1, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD
    for r, row in enumerate(last_run, start=2):
        for i, k in enumerate(headers, start=1):
            ws4.cell(row=r, column=i, value=row[k])
    for col, w in zip("ABCDEFG", [40, 14, 18, 12, 18, 18, 22]):
        ws4.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    with transaction() as conn:
        sync_id = new_id()
        actor = session.get("worker_id")
        conn.execute(
            """INSERT INTO container_excel_sync
               (id, direction, rows_count, bytes, sha256, actor_id)
               VALUES (?, 'export', ?, ?, ?, ?)""",
            (sync_id, len(inflows), len(data),
             hashlib.sha256(data).hexdigest(), actor),
        )
        witness.append(
            conn,
            actor_id=actor,
            service=SERVICE,
            action="excel.exported",
            target_type="excel_sync",
            target_id=sync_id,
            payload={"bytes": len(data), "inflow_rows": len(inflows)},
        )

    fname = f"container_state_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─────────────────────────────────────── single-agent report

@app.get("/agent/report")
def agent_report() -> dict:
    """The one container_agent's read-only summary view.
    Composes container state + milestone health + last scenario for at-a-glance."""
    with transaction() as conn:
        state = _container_state(conn)
        milestones_progress = {
            r["milestone_id"]: _row(r)
            for r in conn.execute("SELECT * FROM container_milestone_progress").fetchall()
        }
        last = conn.execute(
            "SELECT * FROM container_scenario_runs ORDER BY id DESC LIMIT 1"
        ).fetchone()

    today = datetime.now(timezone.utc).date().isoformat()
    overdue = []
    due_soon = []
    done_count = 0
    blocked = []
    for m in MILESTONES:
        p = milestones_progress.get(m["id"], {})
        status = p.get("status", "pending")
        if status == "done":
            done_count += 1
        elif status == "blocked":
            blocked.append(m["id"])
        elif status != "done" and m["date"] < today:
            overdue.append({"id": m["id"], "title": m["title"], "date": m["date"]})
        elif status != "done" and m["date"] <= _add_days(today, 3):
            due_soon.append({"id": m["id"], "title": m["title"], "date": m["date"]})

    health = "ok"
    if state["ahead_or_behind_pct"] < -15 or len(overdue) > 5:
        health = "critical"
    elif state["ahead_or_behind_pct"] < -5 or len(overdue) > 2:
        health = "warning"

    return {
        "today":           today,
        "health":          health,
        "container":       state,
        "milestone_summary": {
            "done":          done_count,
            "total":         len(MILESTONES),
            "overdue":       overdue,
            "due_in_3_days": due_soon,
            "blocked":       blocked,
        },
        "last_scenario_run": _row(last),
    }


def _add_days(date_iso: str, days: int) -> str:
    from datetime import date, timedelta
    d = date.fromisoformat(date_iso)
    return (d + timedelta(days=days)).isoformat()
