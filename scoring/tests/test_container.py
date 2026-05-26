"""Container service tests.

Read order matches the API:
  - frozen target
  - milestones (27, from constants)
  - inflow append-only + milestone link
  - 3 scenarios produce stable, ordered outputs
  - excel round-trip (export -> import)
  - witness log captures every container.* write
"""
import io

import pytest


@pytest.fixture
def cenv(env):
    """env + the Iveel × Dealy seed (container row, members, milestones)."""
    from shared.bootstrap import seed_iveel
    from shared.db import transaction
    with transaction() as conn:
        seed_iveel(conn)
    return env


def test_container_state_starts_at_zero(cenv):
    r = cenv["container"].get("/container")
    assert r.status_code == 200
    j = r.json()
    assert j["target_mnt"] == 1_000_000_000
    assert j["currency"] == "MNT"
    assert j["filled_mnt"] == 0
    assert j["fill_pct"] == 0.0
    assert j["monitor"] is not None
    assert j["monitor"]["handle"] == "container_agent"


def test_milestones_list_is_27_with_critical(cenv):
    r = cenv["container"].get("/milestones")
    assert r.status_code == 200
    j = r.json()
    assert j["total_count"] == 27
    assert j["critical_count"] >= 9
    ids = [m["id"] for m in j["milestones"]]
    assert ids[0] == "M01"
    assert ids[-1] == "M27"
    # Every milestone starts pending.
    assert all(m["status"] == "pending" for m in j["milestones"])
    # All three tracks are represented.
    tracks = {m["track"] for m in j["milestones"]}
    assert tracks == {"tech", "biz", "mkt"}


def test_milestone_status_update_writes_through(cenv):
    r = cenv["container"].post("/milestones/M01/status", json={
        "status": "done", "completion_pct": 100, "notes": "Repo live",
    })
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "done"
    assert body["completion_pct"] == 100
    assert body["completed_at"] is not None

    # And it shows up in the list.
    r = cenv["container"].get("/milestones")
    m = next(m for m in r.json()["milestones"] if m["id"] == "M01")
    assert m["status"] == "done"


def test_unknown_milestone_status_404s(cenv):
    r = cenv["container"].post("/milestones/M99/status", json={
        "status": "done", "completion_pct": 100,
    })
    assert r.status_code == 404


def test_inflow_validates_channel_and_milestone(cenv):
    # Bad channel
    r = cenv["container"].post("/container/inflows", json={
        "occurred_at": "2026-06-05T10:00:00Z",
        "amount_mnt": 100,
        "channel": "not_a_channel",
    })
    assert r.status_code == 422

    # Bad milestone
    r = cenv["container"].post("/container/inflows", json={
        "occurred_at": "2026-06-05T10:00:00Z",
        "amount_mnt": 100,
        "channel": "ecom_site_50off",
        "milestone_id": "M99",
    })
    assert r.status_code == 422


def test_inflow_fills_container_and_milestone(cenv):
    r = cenv["container"].post("/container/inflows", json={
        "occurred_at": "2026-06-05T10:00:00Z",
        "amount_mnt": 80_000_000,
        "channel": "ecom_site_50off",
        "milestone_id": "M09",
        "note": "MVP first week",
    })
    assert r.status_code == 201

    state = cenv["container"].get("/container").json()
    assert state["filled_mnt"] == 80_000_000
    assert state["fill_pct"] == 8.0
    assert state["inflow_count"] == 1

    # Milestone progress tracks the inflow
    ms = cenv["container"].get("/milestones").json()
    m = next(m for m in ms["milestones"] if m["id"] == "M09")
    assert m["actual_revenue_mnt"] == 80_000_000


def test_inflows_pagination_with_uuid7_cursor(cenv):
    # add 5 inflows
    for i in range(5):
        cenv["container"].post("/container/inflows", json={
            "occurred_at": f"2026-06-0{i+1}T10:00:00Z",
            "amount_mnt": 1_000_000 * (i + 1),
            "channel": "ecom_site_50off",
        })
    r = cenv["container"].get("/container/inflows?limit=3")
    assert r.status_code == 200
    page1 = r.json()["inflows"]
    assert len(page1) == 3
    # newest first (UUID7 DESC)
    assert page1[0]["id"] > page1[-1]["id"]


def test_three_scenarios_produce_stable_outputs(cenv):
    expected_goal_ranges = {
        "bear": (50, 75),       # 60-65% target
        "base": (115, 140),
        "bull": (200, 230),
    }
    last_revenue = None
    for key in ["bear", "base", "bull"]:
        r = cenv["container"].post("/scenarios/simulate", json={"scenario_key": key})
        assert r.status_code == 201, r.text
        out = r.json()
        assert out["scenario_key"] == key
        lo, hi = expected_goal_ranges[key]
        assert lo <= out["goal_pct"] <= hi, (key, out["goal_pct"])
        if last_revenue is not None:
            # bull > base > bear
            assert out["total_revenue_mnt"] > last_revenue
        last_revenue = out["total_revenue_mnt"]
        # Win-win only on base / bull (bear A-side goes negative)
        if key == "bear":
            assert out["win_win"] is False
        else:
            assert out["win_win"] is True


def test_scenario_runs_persist(cenv):
    for key in ["bear", "base", "bull"]:
        cenv["container"].post("/scenarios/simulate", json={"scenario_key": key})
    r = cenv["container"].get("/scenarios/runs")
    runs = r.json()["runs"]
    assert len(runs) == 3
    assert {r["scenario_key"] for r in runs} == {"bear", "base", "bull"}


def test_excel_export_returns_xlsx(cenv):
    r = cenv["container"].get("/excel/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 1000  # actual workbook bytes
    # Verify the workbook is valid by re-opening it.
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"Container", "Milestones", "Inflows", "Scenario_Runs"}
    # Milestones sheet has 27 + header rows
    ws = wb["Milestones"]
    assert ws.max_row == 28


def test_excel_import_round_trip(cenv):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["occurred_at", "amount_mnt", "channel", "milestone_id", "note"])
    ws.append(["2026-06-10T15:00:00Z", 25_000_000, "fb_messenger_chatbot", "M14", "From chat"])
    ws.append(["2026-06-15T12:00:00Z", 15_000_000, "tourist_channel",      "M17", "Tour #1"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {"file": ("inflows.xlsx", buf.read(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = cenv["container"].post("/excel/import", files=files)
    assert r.status_code == 200
    j = r.json()
    assert j["imported"] == 2
    assert j["errors"] == []

    state = cenv["container"].get("/container").json()
    assert state["filled_mnt"] == 40_000_000


def test_excel_import_reports_row_level_errors(cenv):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["occurred_at", "amount_mnt", "channel"])
    ws.append(["2026-06-10T15:00:00Z", 1_000_000, "ecom_site_50off"])   # valid
    ws.append(["2026-06-10T15:00:00Z", 1_000_000, "not_a_channel"])      # invalid
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {"file": ("mixed.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = cenv["container"].post("/excel/import", files=files)
    assert r.status_code == 200
    j = r.json()
    assert j["imported"] == 1
    assert len(j["errors"]) == 1


def test_agent_report_health_signal(cenv):
    r = cenv["container"].get("/agent/report")
    assert r.status_code == 200
    rep = r.json()
    assert rep["health"] in {"ok", "warning", "critical"}
    assert rep["milestone_summary"]["total"] == 27
    assert rep["milestone_summary"]["done"] == 0


def test_witness_log_records_container_writes(cenv):
    # Do a few writes
    cenv["container"].post("/milestones/M01/status", json={"status": "done", "completion_pct": 100})
    cenv["container"].post("/container/inflows", json={
        "occurred_at": "2026-06-05T10:00:00Z",
        "amount_mnt": 50_000_000, "channel": "ecom_site_50off",
    })
    cenv["container"].post("/scenarios/simulate", json={"scenario_key": "base"})

    import sqlite3
    conn = sqlite3.connect(cenv["db_path"])
    conn.row_factory = sqlite3.Row
    actions = [r["action"] for r in conn.execute(
        "SELECT action FROM judge_witness_log WHERE service='container' ORDER BY id"
    )]
    conn.close()
    assert "milestone.status" in actions
    assert "inflow.recorded" in actions
    assert "scenario.simulated" in actions


def test_inflows_require_auth(cenv):
    r = cenv["guest_container"].post("/container/inflows", json={
        "occurred_at": "2026-06-05T10:00:00Z",
        "amount_mnt": 1_000_000,
        "channel": "ecom_site_50off",
    })
    assert r.status_code == 401


def test_excel_import_requires_admin(cenv):
    files = {"file": ("x.xlsx", b"\x50\x4b\x03\x04", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = cenv["guest_container"].post("/excel/import", files=files)
    assert r.status_code == 401


def test_constants_immutable_signature(cenv):
    """The frozen plan should match constants exactly. If it doesn't, the seed
    drifted from the source of truth and we want loud test failure."""
    from shared.constants import MILESTONES, CONTAINER_TARGET, total_expected_revenue
    assert CONTAINER_TARGET == 1_000_000_000
    assert len(MILESTONES) == 27
    # Order is the source of truth too.
    assert MILESTONES[0]["id"] == "M01"
    assert MILESTONES[-1]["id"] == "M27"
    # Sum of per-milestone expected revenue is what we calibrated.
    assert total_expected_revenue() > 800_000_000
